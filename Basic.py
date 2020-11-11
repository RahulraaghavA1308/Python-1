'''
a =input()    #if else explanation
b=input()
if a>b:
    print(a)
    print("is greater")
elif  b>a:
    print(b +"is greater")
else :
    print("No result")
'''
'''
if 1 and 0:  #and ,or  operators
    print("true")
elif 1 or 0:
    print("very true")
'''
'''
a=int(input())  #comparing two numbers
b=int(input())
print(a)
if a==b:
    print("same")
else:
    print("diff")
'''
'''
print("enter your weight")   #converting Kg to pound
a=int(input())
print("enter unit")
b=str(input())
if b=='k':
    print(.45*a)
elif b=='l':
    print(a/.45)
'''
'''
for item in range(5,10):             #printing numbers using range from 5 to 10
    print(item*10)
for item in range(5,10,2):          #printing numbers using range from 5 to 10 with a step of 2
    print(item*10)
'''
'''
item=[2,5,7,3,5,6]              #for finding total sum of elements in a list
total=0
for i in item:
    total+=i
print(total)
'''
'''
for x in range(5):  
    print(x)
    for y in range(6):
        print(y)
'''
'''list =[5,2,5,2,2]
c=input()
for i in list:
    outtp=" "
    for j in range(i):
        outtp += "x"
    print(outtp)
 '''
'''                       #usage of lists
names =['rahul','ajay','idfoo']  
names[0]='rocky'
print(names[0:3])
'''
'''                 #finding largest number in a list 
num = [100000,3,4,2,5,23,54,2323]
max=num[0]
for i in num:
    if max<i:
        max=i
print(max)
'''
'''           ## 2D  LISTS
matrix =[[1,2,4],[4,6,4],[39,6,2,4]]
matrix[0][2]=5
#print(matrix[0][2])

for i in matrix:
    for j in i:
        print(j)
'''
'''
num=[5,1,2,3,4,5,20]
num.append(25)
print(num)
num.insert(0,100)    #to insert an element passed at specified index
print(num)
num.remove(3)
print(num)
print(num.pop())        #to pop an element
print(num)
print(num.index(5))    #to find index of a object that is passed
print(2 in num)  #to check existence of a num in a list
print(num.count(5))   #to pront no of ocurrences of passed value
num_2=num.copy()        #to make an copy of the list
num.sort()              #to sort an list
print(num)
print(num_2)
'''
'''num=[5,1,2,3,4,4,5,6,5,20]
num.remove(4)
print(num)

'''
'''             #For removing duplicates  in a list
num=[2,5,3,2,4,1,3,4,2,4]
num_2=[ ]
for i in num:
    if i not in num_2:
        num_2.insert(0,i)
num_2.sort()
print(num_2)
'''
'''        #tuples are immutable
num=(1,2,3,2,5,3,4)
print(num.count(3))
print(num[1])
'''
'''            ##UNPACKING
count =(1,2,2,4,2,82)
a,b,c,d,e,f =count   #a,b,...  hold the values of every element in count
print(b)
'''
''' ###DICTIONARY
customer ={
    "name" :"rahul",  #these are key value pairs
                      #here name ,age,mail are the keys
    "age" :"23",      #keys should be unique
    "mail":"rahuj.com",
    "is_verified": True
}
customer["DOB"] ="aug 2002"         #we are adding a new key
customer["name"] ="rahul raaghav"
print(customer["name"])
print(customer.get("DOB","jan 2002"))  #if key doesnt exist we can pass a default value
A=customer.get("name")
print(A)'''
'''
''''''                #FUNCTIONS
def wel_message(A1, A2):                     # A is the parameter being passed
    print("hello there")
    print("welcome to my program")
    print(f'{A1} {A2}')                 #using f   to  format strings
                                        #two blank lines are required to
                                        # end the def of a Func
a=input()
b=input()
wel_message(a,b)             #func call   a is the argument
wel_message(A2=b,A1=a)       #key word arguments
'''
            #return statements
            #FIBONCCI generation
''''''
'''def fibo(n):
    f = -1
    s = 1
    for i in range(n + 1):
        t = f + s
        f = s
        s = t
    return t


while True:
    print(fibo(int(input())))
    if input() !='y':
        break
'''
                ##exception
                ##used to predict errors
                ##to prevent errors from crashing the program
'''try:
    age = int(input("enter your name :"))
    print(age)
except ValueError:
    print("invalid name")

try:
    age = int(input("enter a  no  :"))
    c = 233
    print(c/age)
except ZeroDivisionError:     #to solve division by zero problem
    print("Division by zero not possible")
'''
            ####CLASSES
'''
class Point:
    def move(self):
        print("move")
    def draw(self):
        print("fu")


a = Point()             ##creating an object of a class
a.draw()
a.x=12
print(a.x)
'''
                        #Classes
'''
class Point:                                #self is an reference to the calling obeject itself
    def __init__(self, a=0, b=0):           #default argument
        self.x=a                            #Constructors
        self.y=b
    def move(self):
        print("move")
    def draw(self):
        print(f"hi you  chose {self.x}")


A = Point(int(input()),int(input()))

A.draw()
print(A.x)
print(A.y)
'''
                        #Inheritance
'''
class Mammal:
    def walk(self):
        print("gdg")


class Dog(Mammal):
    pass


class Cat(Mammal):
    pass


dog1 = Dog()
dog1.walk()
'''
                    #Modules
'''
import fibonaci             #importing  function fibo from another file

#from fibonaci import fibo
#print(fibo(int(input())))

#print(fibonaci.fibo(int(input())))

A =[1213132,1232,412312,112313233,5123,2213,4123,25345,5323,5223323,231233,]
print(fibonaci.max(A))
'''
'''import random
for i in range(10):
    print(random.randint(0,10))
'''
#or
'''
from random import choice
A=("rahul","Ajay","Vijay")
for i in range(3):
    print(choice(A))
'''
'''
            #assignment
            #Rolling an Dice
from random import randint
class Dice:
    def roll(self):
        print(randint(1,6))
        print(randint(1,6))


A = Dice()
A.roll()
'''

            #Files and Directries
''''from pathlib import Path
path=Path("project 2020 Django")
#print(path.exists())
#path.mkdir()           ##for creating an directory named path
path.rmdir()            ##for deleting an directory named path


path2=Path()
for file in path2.glob("*"): 
    print(file)
'''
            ##pypi  and pip
    ##selena package

##working with spread sheets
'''import openpyxl as xl  ##alias name
from openpyxl.chart import  BarChart,Reference      #for drawing bar graphs

wb =xl.load_workbook("transactions.xlsx")       #Open the given filename and return the workbook
sheet=wb['Sheet1']  #storing in a sheet
cell =sheet['a1']   #accesing Cells
cell2=sheet.cell(1, 2)          #a1 refers to col 1 row1
print(cell.value)           #prints value of cell
print(cell2.value)
for row in range(2, sheet.max_row + 1):
    cell3 = sheet.cell(row,3)
    print(cell3.value)
    alter=cell3.value *0.9
    print(alter)
    alter_cell =sheet.cell(row,4)
    alter_cell.value =alter
#wb.save('transactions2.xlsx')           #Save the current workbook under the given filename.
# #Use this function instead of using an ExcelWriter.
#new file will be created at Current directory

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'f3')
wb.save('transactions3.xlsx')
#print("")
#for r in range(2,sheet.max_row+1):
#    print(sheet.cell(r,4).value)
'''

